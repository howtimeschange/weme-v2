"""
批量发送任务模块
--------------
数据模型 + Excel 解析 + 模版生成
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Optional


# ── 枚举 ─────────────────────────────────────────────────────────────────────

class TaskStatus(str, Enum):
    PENDING   = "待发送"
    SCHEDULED = "已计划"
    RUNNING   = "发送中"
    DONE      = "已完成"
    FAILED    = "失败"
    SKIPPED   = "跳过"


class MsgType(str, Enum):
    TEXT  = "文字"
    IMAGE = "图片"
    BOTH  = "文字+图片"


class AppTarget(str, Enum):
    WECHAT   = "wechat"
    DINGTALK = "dingtalk"
    FEISHU   = "feishu"

    @classmethod
    def from_str(cls, s: str) -> "AppTarget":
        mapping = {
            "微信": cls.WECHAT, "wechat": cls.WECHAT,
            "钉钉": cls.DINGTALK, "dingtalk": cls.DINGTALK,
            "飞书": cls.FEISHU, "lark": cls.FEISHU, "feishu": cls.FEISHU,
        }
        return mapping.get(s.strip().lower(), cls.WECHAT)


# ── 数据模型 ──────────────────────────────────────────────────────────────────

@dataclass
class SendTask:
    """一条发送任务"""

    # 基础信息（对应 Excel 列）
    row_num: int               # Excel 行号（从 2 开始）
    app: AppTarget             # 目标应用
    target: str                # 联系人/群聊名称
    msg_type: MsgType          # 消息类型
    text: str                  # 文字内容（可含变量 {name} 等）
    image_path: str            # 图片路径（msg_type 含图片时必填）
    send_at: Optional[datetime]# 定时发送时间；None = 立即发送
    repeat: str                # 重复规则："" / "daily" / "weekly" / "workday"
    comment: str               # 备注（不参与发送）

    # 运行时
    status: TaskStatus = TaskStatus.PENDING
    error: str = ""
    sent_at: Optional[datetime] = None

    @property
    def is_immediate(self) -> bool:
        return self.send_at is None

    @property
    def is_scheduled(self) -> bool:
        return self.send_at is not None

    def render_text(self, variables: dict[str, str] | None = None) -> str:
        """渲染文字内容，替换 {变量名}"""
        if not variables:
            return self.text
        result = self.text
        for k, v in variables.items():
            result = result.replace(f"{{{k}}}", v)
        return result


# ── Excel 列定义 ──────────────────────────────────────────────────────────────

# 列顺序和 header 名称（严格匹配）
COLUMNS = [
    ("row_marker",  "#"),
    ("app",         "应用"),           # 微信 / 钉钉 / 飞书
    ("target",      "联系人/群聊"),
    ("msg_type",    "消息类型"),        # 文字 / 图片 / 文字+图片
    ("text",        "文字内容"),
    ("image_path",  "图片路径"),
    ("send_at",     "发送时间"),        # 留空=立即；格式 2026-03-24 14:30
    ("repeat",      "重复"),            # 留空 / daily / weekly / workday
    ("comment",     "备注"),
    ("status",      "状态"),            # 由程序写入，业务无需填
]

COL_LETTERS = {key: chr(65 + i) for i, (key, _) in enumerate(COLUMNS)}  # A B C ...
HEADER_ROW = {name: key for key, name in COLUMNS}


# ── Excel 生成 ────────────────────────────────────────────────────────────────

def create_template(output_path: Path) -> Path:
    """生成 Excel 模版文件，带样式、下拉校验和示例行"""
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发送任务"

    # ── 样式常量 ──────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1A1D27")
    ACCENT_FILL  = PatternFill("solid", fgColor="6C8CFF")
    ALT_FILL     = PatternFill("solid", fgColor="F4F5F9")
    DONE_FILL    = PatternFill("solid", fgColor="D4EDDA")
    FAIL_FILL    = PatternFill("solid", fgColor="F8D7DA")

    HEADER_FONT  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    BODY_FONT    = Font(name="微软雅黑", size=10)
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── 主标题行 ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "🦐 Weme 批量发送任务表  |  填写说明：带 * 号列为必填；发送时间留空 = 立即发送"
    title_cell.font  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=12)
    title_cell.fill  = PatternFill("solid", fgColor="3D50CC")
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # ── 列标题行 ──────────────────────────────────────────────────────────────
    headers = [name for _, name in COLUMNS]
    required = {"应用", "联系人/群聊", "消息类型", "文字内容"}

    for col_idx, (key, name) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = f"* {name}" if name in required else name
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 24

    # ── 列宽 ─────────────────────────────────────────────────────────────────
    col_widths = [5, 10, 20, 14, 40, 28, 20, 12, 20, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 数据校验（下拉）──────────────────────────────────────────────────────
    dv_app = DataValidation(
        type="list",
        formula1='"微信,钉钉,飞书"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="输入错误",
        error='请从下拉列表选择：微信 / 钉钉 / 飞书',
    )
    dv_type = DataValidation(
        type="list",
        formula1='"文字,图片,文字+图片"',
        allow_blank=False,
    )
    dv_repeat = DataValidation(
        type="list",
        formula1='"留空=单次,daily,weekly,workday"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_app)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_repeat)
    dv_app.add(f"B3:B1000")
    dv_type.add(f"D3:D1000")
    dv_repeat.add(f"H3:H1000")

    # ── 示例行 ───────────────────────────────────────────────────────────────
    examples = [
        # #  app    target           type          text                         image  time           repeat  comment  status
        [1, "微信", "产品讨论群",    "文字",       "大家好，今天下午3点开会，请准时参加。", "",    "",              "",       "例：群发通知",   ""],
        [2, "微信", "张三",          "文字+图片",  "这是今日的周报，请查收。",             "/Users/you/report.png", "2026-03-24 15:00", "", "例：定时发周报", ""],
        [3, "飞书", "研发部-全员群", "文字",       "⚠️ 系统将于今晚 22:00 维护，请提前保存工作。", "", "2026-03-24 21:45", "workday", "例：定时+重复",  ""],
        [4, "钉钉", "李四",          "图片",       "",                                    "/Users/you/banner.jpg", "",   "",       "例：只发图片",   ""],
    ]
    for row_idx, row_data in enumerate(examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = LEFT if col_idx in (3, 5, 6, 9) else CENTER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
        ws.row_dimensions[row_idx].height = 20

    # ── 说明 Sheet ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ("字段",       "说明",                                       "示例"),
        ("应用",       "微信 / 钉钉 / 飞书（三选一）",               "微信"),
        ("联系人/群聊","精确的联系人昵称或群聊名称（用于搜索）",      "产品讨论群"),
        ("消息类型",   "文字 / 图片 / 文字+图片",                    "文字+图片"),
        ("文字内容",   "支持变量：{name} {date} {time}",             "你好 {name}，今日报告请查收"),
        ("图片路径",   "本机绝对路径，支持 jpg/png/gif",             "/Users/you/pic.jpg"),
        ("发送时间",   "留空=立即发送；格式 YYYY-MM-DD HH:MM",       "2026-03-24 15:00"),
        ("重复",       "留空=单次；daily=每天；weekly=每周；workday=工作日", "daily"),
        ("备注",       "仅供人类阅读，不影响发送",                   "Q1 通知"),
        ("状态",       "由程序自动填入，无需手动填写",               "已完成"),
    ]
    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(name="微软雅黑", bold=True, size=11)
                cell.fill = PatternFill("solid", fgColor="3D50CC")
                cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
            else:
                cell.font = Font(name="微软雅黑", size=10)
                if r % 2 == 0:
                    cell.fill = ALT_FILL
            cell.alignment = LEFT
            cell.border = BORDER

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 28

    wb.save(output_path)
    return output_path


# ── Excel 解析 ────────────────────────────────────────────────────────────────

_DT_FORMATS = [
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%m/%d %H:%M",
    "%m-%d %H:%M",
]


def _parse_datetime(raw) -> Optional[datetime]:
    if raw is None or str(raw).strip() in ("", "留空=单次"):
        return None
    if isinstance(raw, datetime):
        return raw
    s = str(raw).strip()
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _parse_repeat(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    if s in ("", "留空=单次"):
        return ""
    if s in ("daily", "weekly", "workday"):
        return s
    return ""


def parse_excel(path: Path) -> list[SendTask]:
    """解析 Excel 文件，返回 SendTask 列表（跳过空行和标题）。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    tasks: list[SendTask] = []
    header_row_idx = None

    # 找标题行（含「联系人/群聊」的行）
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        row_texts = [str(c or "").strip().lstrip("* ") for c in row]
        if "联系人/群聊" in row_texts or "target" in row_texts:
            header_row_idx = ws.iter_rows().__next__  # 记录下标
            break

    # 按列名映射（找到 header 行后跳过）
    start_row = None
    col_map: dict[str, int] = {}  # key -> 0-based column index

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        row_texts = [str(c or "").strip().lstrip("* ") for c in row]
        if "联系人/群聊" in row_texts:
            start_row = row_idx + 1
            for col_idx, txt in enumerate(row_texts):
                for key, name in COLUMNS:
                    if txt == name:
                        col_map[key] = col_idx
            break

    if not col_map or start_row is None:
        raise ValueError("找不到标题行，请确认使用 weme 生成的标准模版")

    def get(row_vals, key, default=""):
        idx = col_map.get(key)
        if idx is None or idx >= len(row_vals):
            return default
        v = row_vals[idx]
        return v if v is not None else default

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=start_row, values_only=True), start_row
    ):
        # 跳过完全空行
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        target = str(get(row, "target", "")).strip()
        if not target:
            continue

        app_str = str(get(row, "app", "微信")).strip()
        type_str = str(get(row, "msg_type", "文字")).strip()
        send_at_raw = get(row, "send_at", None)
        repeat_raw = get(row, "repeat", "")

        try:
            msg_type = {
                "文字": MsgType.TEXT,
                "图片": MsgType.IMAGE,
                "文字+图片": MsgType.BOTH,
            }.get(type_str, MsgType.TEXT)

            task = SendTask(
                row_num=row_idx,
                app=AppTarget.from_str(app_str),
                target=target,
                msg_type=msg_type,
                text=str(get(row, "text", "")).strip(),
                image_path=str(get(row, "image_path", "")).strip(),
                send_at=_parse_datetime(send_at_raw),
                repeat=_parse_repeat(repeat_raw),
                comment=str(get(row, "comment", "")).strip(),
                status=TaskStatus.PENDING,
            )
            tasks.append(task)
        except Exception as exc:
            # 记录解析失败但不中断
            tasks.append(SendTask(
                row_num=row_idx,
                app=AppTarget.WECHAT,
                target=target or f"row{row_idx}",
                msg_type=MsgType.TEXT,
                text="",
                image_path="",
                send_at=None,
                repeat="",
                comment="",
                status=TaskStatus.FAILED,
                error=f"解析错误: {exc}",
            ))

    return tasks


def write_status_back(path: Path, tasks: list[SendTask]) -> None:
    """将任务状态回写到 Excel 的「状态」列"""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    status_col = None
    start_data_row = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_texts = [str(c or "").strip().lstrip("* ") for c in row]
        if "状态" in row_texts or "status" in row_texts:
            for col_idx, txt in enumerate(row_texts):
                if txt in ("状态", "status"):
                    status_col = col_idx + 1  # 1-based
            start_data_row = row_idx + 1
            break

    if status_col is None:
        return

    from openpyxl.styles import PatternFill, Font
    STATUS_COLORS = {
        TaskStatus.DONE:      ("D4EDDA", "155724"),
        TaskStatus.FAILED:    ("F8D7DA", "721C24"),
        TaskStatus.RUNNING:   ("D1ECF1", "0C5460"),
        TaskStatus.SCHEDULED: ("FFF3CD", "856404"),
        TaskStatus.PENDING:   ("FFFFFF", "333333"),
        TaskStatus.SKIPPED:   ("E2E3E5", "383D41"),
    }

    row_map = {t.row_num: t for t in tasks}
    for task in tasks:
        row_cell = ws.cell(row=task.row_num, column=status_col)
        row_cell.value = task.status.value
        if task.error:
            row_cell.value += f": {task.error[:30]}"
        bg, fg = STATUS_COLORS.get(task.status, ("FFFFFF", "333333"))
        row_cell.fill = PatternFill("solid", fgColor=bg)
        row_cell.font = Font(name="微软雅黑", color=fg, size=10)

    wb.save(path)
